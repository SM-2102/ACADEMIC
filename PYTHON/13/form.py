# data entry and load data

import tkinter
from tkinter import messagebox,ttk
import os
import openpyxl
from tkcalendar import DateEntry

def entry():

    def enter_data():
        accept = accept_var.get()

        if accept == 1:
            firstname = first_name_entry.get()
            middlename = middle_name_entry.get()
            lastname = last_name_entry.get()
            
            if firstname and lastname:
                if firstname.isalpha() and lastname.isalpha():
                    if middlename:
                        middlename = middlename + ' '
                    name = firstname+' '+middlename+lastname
                    name = name.upper()
                    age = age_spinbox.get()
                    dateofbirth = dob_entry.get()
                    gender =gen()
                    contactno=contactno_entry.get()
                    email = email_entry.get()

                    department = dept.get()
                    if department not in departments:
                        messagebox.showwarning(title="Error", message="Enter your department.")

                    semester = sem_spinbox.get()
                    rollno = rollno_entry.get()

                    if len(str(contactno))==10:
                        try:
                            contactno = int(contactno)
                        except:
                            messagebox.showwarning(title="Error", message="Enter a contact number.")
                    else:
                        messagebox.showwarning(title="Error", message="Enter correct contact number.")
                    
                    if not os.path.exists(path):
                        workbook = openpyxl.Workbook()
                        sheet = workbook.active
                        heading = ['Name','Age','Gender','Contact No','Department','Semester','Roll No','Date of Birth','Email']
                        sheet.append(heading)
                        workbook.save(path)
                    workbook = openpyxl.load_workbook(path)
                    sheet = workbook.active
                    sheet.append([name, age, gender, contactno,department,semester,rollno,dateofbirth,email]) 
                    workbook.save(path)
                else:
                    messagebox.showwarning(title="Error", message="Enter proper name.")      
            else:
                messagebox.showwarning(title="Error", message="First name and last name are required.")
        else:
            messagebox.showwarning(title= "Error", message="You have not accepted the terms")

    window = tkinter.Tk()
    window.title("Data Entry Form")

    frame = tkinter.Frame(window)
    frame.pack()
    
    def gen():
        gen = var.get()
        if gen==1:
            gender = 'Male'
        elif gen==2:
            gender = 'Female'
        else:
            gender = 'Others'
        return gender

    pers_info =tkinter.LabelFrame(frame, text="Personal Information")
    pers_info.grid(row= 0, column=0, padx=20, pady=10)

    first_name_label = tkinter.Label(pers_info, text="First Name")
    first_name_label.grid(row=0, column=0)
    middle_name_label = tkinter.Label(pers_info, text="Middle Name")
    middle_name_label.grid(row=0, column=1)
    last_name_label = tkinter.Label(pers_info, text="Last Name")
    last_name_label.grid(row=0, column=2)

    first_name_entry = tkinter.Entry(pers_info)
    first_name_entry.grid(row=1, column=0)
    middle_name_entry = tkinter.Entry(pers_info)
    middle_name_entry.grid(row=1, column=1)
    last_name_entry = tkinter.Entry(pers_info)
    last_name_entry.grid(row=1, column=2)

    age_label = tkinter.Label(pers_info, text="Age")
    age_spinbox = tkinter.Spinbox(pers_info, from_=18, to=110)
    age_label.grid(row=2, column=0)
    age_spinbox.grid(row=3, column=0)

    dob_label = tkinter.Label(pers_info, text="Date of Birth")
    dob_label.grid(row=2, column=1)
    dob_entry=DateEntry(pers_info,selectmode='day',date_pattern='dd-MM-yyyy')
    dob_entry.grid(row=3,column=1)

    gender_label = tkinter.Label(pers_info, text="Gender")
    gender_label.grid(row=2, column=2)
    var = tkinter.IntVar(pers_info)
    radiobutton1 = tkinter.Radiobutton(pers_info, variable=var,text="Male", value=1,command=gen)
    radiobutton1.grid(row=3, column=2)
    radiobutton2 = tkinter.Radiobutton(pers_info, variable=var,text="Female", value=2,command=gen)
    radiobutton2.grid(row=4, column=2)
    radiobutton3 = tkinter.Radiobutton(pers_info, variable=var,text="Others", value=3,command=gen)
    radiobutton3.grid(row=5, column=2)

    contactno_label = tkinter.Label(pers_info, text="Contact Number")
    contactno_label.grid(row=4, column=0)
    contactno_entry = tkinter.Entry(pers_info)
    contactno_entry.grid(row=5, column=0)

    email_label = tkinter.Label(pers_info, text="Email Id")
    email_label.grid(row=4, column=1)
    email_entry = tkinter.Entry(pers_info)
    email_entry.grid(row=5, column=1)

    for widget in pers_info.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    acad_info = tkinter.LabelFrame(frame,text = 'Academic Information')
    acad_info.grid(row=1, column=0, padx=20, pady=10)

    departments = ['CSE','IT','ECE','EE','AIML']
    dept_label = tkinter.Label(acad_info, text="Department")
    dept_label.grid(row=0, column=0)
    dept = ttk.Combobox(acad_info, values=departments)
    dept.grid(row=1,column=0)

    sem_label = tkinter.Label(acad_info, text="Semester")
    sem_label.grid(row=0, column=1)
    sem_spinbox = tkinter.Spinbox(acad_info, from_=1, to=8)
    sem_spinbox.grid(row=1, column=1)

    rollno_label = tkinter.Label(acad_info, text="Roll Number")
    rollno_label.grid(row=0, column=2)
    rollno_entry = tkinter.Entry(acad_info)
    rollno_entry.grid(row=1, column=2)

    for widget in acad_info.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    terms_frame = tkinter.LabelFrame(frame, text="Terms & Conditions")
    terms_frame.grid(row=2, column=0, sticky="news", padx=20, pady=10)  

    accept_var = tkinter.IntVar(terms_frame)
    terms_check = tkinter.Checkbutton(terms_frame, text= "I accept the terms and conditions.", 
                                        variable=accept_var, onvalue=1, offvalue=0)
    terms_check.grid(row=0, column=0)

    button = tkinter.Button(frame, text="Submit", command= enter_data,activebackground='#E6E6FA')
    button.grid(row=3, column=0, padx=20, pady=10)
    
    window.mainloop()


def load():

    if not os.path.exists(path):
        messagebox.showwarning(title="Error", message="File Not Found.")
        return
    
    window = tkinter.Tk()
    window.title("List of Entries")

    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    list_values = list(sheet.values)

    if len(list_values)==1:
        messagebox.showwarning(title="Error", message="No Entries Found.")
        window.destroy()

    cols = list_values[0]
    tree = ttk.Treeview(window, column= cols, show="headings")
    widths = [150,80,70,90,120,70,80,100,250]
    i=0
    for col_name in cols:
        tree.heading(col_name, text = col_name)
        tree.column(col_name,anchor="center",width=widths[i])
        i=i+1
    tree.pack(expand=True)
        
    for value in list_values[1:]:
        tree.insert('',tkinter.END, values=value)

    window.mainloop()

def search():
    
    if not os.path.exists(path):
        messagebox.showwarning(title="Error", message="File Not Found.")
        return

    window = tkinter.Tk()
    window.title("Search Database")
    
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    list_values = list(sheet.values)
    head = list_values[0]
    list_values.pop(0)
 
    if not list_values:
        messagebox.showerror(title="Error", message="No Entries Found")
        window.destroy()

    def search_data():
        category = categ.get()
        data = data_entered.get()
        if category == 'Name' or category=='Department':
            data=data.upper()
        elif category == 'Gender':
            data = data.capitalize()
        col = categories.index(category)
        found = []
        
        for i in range(0,len(list_values)):
            if list_values[i][col] == data:
                found.extend([list_values[i]])

        if not found:
            messagebox.showerror(title="Error", message="No Entries Found")
        else:
            window1 = tkinter.Tk()
            window1.title("Searched Results")

            tree = ttk.Treeview(window1, column= head, show="headings")
            widths = [150,80,70,90,120,70,80,100,250]
            i=0
            for col_name in head:
                tree.heading(col_name, text = col_name)
                tree.column(col_name,anchor="center",width=widths[i])
                i=i+1
            tree.pack(expand=True)
                
            for value in found:
                tree.insert('',tkinter.END, values=value)
            window1.mainloop()
        
    options = tkinter.Label(window)
    options.grid(row=0, column=0,pady=10)
    category_label = tkinter.Label(options, text="Category")
    category_label.grid(row=0, column=0)
    enter_label = tkinter.Label(options, text="Enter Data :")
    enter_label.grid(row=0, column=1)
    categories = ['Name','Age','Gender','Contact No','Department','Semester','Roll No']
    categ = ttk.Combobox(options, values=categories)
    categ.grid(row=1,column=0,padx=10,pady=10)
    data_entered = tkinter.Entry(options,width=30)
    data_entered.grid(row=1, column=1,padx=10,pady=10)
    btn_search = tkinter.Button(window, text="Search", width=10,activebackground = '#E6E6FA',justify='center',command=search_data)
    btn_search.grid(row=1, column=0,pady=10,padx=20)

    window.mainloop()

def close():
    root.destroy()

if __name__ == "__main__":
    root = tkinter.Tk()
    root.title('Data Collection')  

    path = "D:\data.xlsx"

    label = tkinter.Label(root, text="STUDENTS' RECORD FOR 2022 - 2023",font = ('Lucida Console',22))
    label.grid(row=0, column=0,pady=15,padx=10)
    options = tkinter.Label(root)
    options.grid(row=1, column=0)
    btn_1 = tkinter.Button(options, text="ENTER DATA", height=2, width=15,activebackground = '#E6E6FA',command = entry)
    btn_1.grid(row=1, column=0,pady=10,padx=20)
    btn_2 = tkinter.Button(options, text="LOAD DATA", height=2, width=15,activebackground = '#90EE90',command = load)
    btn_2.grid(row=1, column=1,pady=10,padx=20)
    btn_3 = tkinter.Button(options, text="SEARCH DATA", height=2, width=15,activebackground = '#ADD8E6',command=search)
    btn_3.grid(row=1, column=2,pady=10,padx=20)
    btn_4 = tkinter.Button(options, text="EXIT", height=2, width=15,activebackground = '#FFCCCB',command=close)
    btn_4.grid(row=1, column=3,pady=10,padx=20)

    root.mainloop()